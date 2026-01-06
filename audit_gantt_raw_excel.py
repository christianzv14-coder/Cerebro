
import openpyxl

def audit_gantt_raw():
    fpath = "outputs/gantt_FINAL_RAW_314.xlsx"
    print(f"Auditing Raw File: {fpath}")
    
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    
    # Raw Layout
    COL_CITY = 1  # A
    COL_TOTAL = 3 # C
    COL_START_DAY = 4 # D
    START_ROW = 5
    
    grand_total_col = 0
    grand_total_grid = 0
    
    row_count = 0
    
    for r in range(START_ROW, 500):
        # Check if row has City
        city_val = ws.cell(row=r, column=COL_CITY).value
        
        # Stop if we hit "TOTAL VERIFICADO" or empty
        if not city_val:
            continue
        if str(city_val).startswith("TOTAL VERIFICADO"):
            print("Hit Verification Row. Stopping scan.")
            break
            
        row_count += 1
        
        # Sum Column Total
        t_val = ws.cell(row=r, column=COL_TOTAL).value
        if t_val:
            try:
                grand_total_col += float(t_val)
            except: pass
            
        # Sum Grid
        row_grid_sum = 0
        for c in range(COL_START_DAY, COL_START_DAY + 62):
            d_val = ws.cell(row=r, column=c).value
            if d_val is not None:
                try:
                    val = float(d_val)
                    row_grid_sum += val
                    grand_total_grid += val
                except:
                    pass
        
        # Mismatch per row?
        if t_val and abs(float(t_val) - row_grid_sum) > 0.01:
            print(f"Row {r} Mismatch: Col Total={t_val} vs Grid Sum={row_grid_sum} (City: {city_val})")

    print(f"--- AUDIT RESULTS ---")
    print(f"Rows Found: {row_count}")
    print(f"Sum of 'Total' Column: {grand_total_col}")
    print(f"Sum of 'Daily Grid': {grand_total_grid}")
    
    if abs(grand_total_col - 314) < 0.1 and abs(grand_total_grid - 314) < 0.1:
        print("SUCCESS: Excel verifies to 314.")
    else:
        print("FAILURE: Excel does NOT sum to 314.")

if __name__ == "__main__":
    audit_gantt_raw()
