import openpyxl
import json
from openpyxl.styles import PatternFill, Alignment, Border, Side

def generate_gantt():
    print("[INFO] Generating ENEX Gantt Chart...")
    
    input_json = "outputs/vrp_result_enex.json"
    template_file = "temp_gantt.xlsx"
    output_file = "outputs/carta_gantt_ENEX.xlsx"
    
    try:
        with open(input_json, "r") as f:
            data = json.load(f)
        plan = data['plan']
    except FileNotFoundError:
        print(f"[ERROR] Could not find {input_json}. Run VRP first.")
        return

    # Group by (City, Tech)
    grouped = {}
    for item in plan:
        c = item['city']
        t = item.get('tech', 'External')
        key = (c, t)
        if key not in grouped:
            grouped[key] = {'total': 0, 'days': {}}
            
        grouped[key]['total'] += item['gps']
        d = item['day']
        grouped[key]['days'][d] = grouped[key]['days'].get(d, 0) + item['gps']
        
    sorted_keys = sorted(grouped.keys(), key=lambda x: (x[0], x[1])) # Sort by City
    
    # Load Template
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active 
    
    # Constants
    COL_CITY = 4 # D
    COL_TOTAL = 5 # E
    COL_TECH = 6 # F
    COL_START_DAY = 7 # G
    START_ROW = 5
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Tech Colors (Mapped for ENEX)
    colors = {
        "Carlos": "FFC7CE", "Luis": "C6EFCE", "Orlando": "FFEB9C",
        "Pedro": "E2EFDA", "Jimmy": "EDEDED", "Wilmer": "D9E1F2",
        "Carlos (Sur)": "FFC7CE", "Pedro (Norte)": "E2EFDA"
    }

    # Clear Old Data
    ws.delete_rows(START_ROW, 500)
    
    curr = START_ROW
    total_gps_verify = 0
    
    for (city, tech) in sorted_keys:
        gdata = grouped[(city, tech)]
        total_gps_verify += gdata['total']
        
        # Metadata Columns
        ws.cell(row=curr, column=COL_CITY, value=city).border = thin_border
        ws.cell(row=curr, column=COL_TOTAL, value=gdata['total']).border = thin_border
        ws.cell(row=curr, column=COL_TECH, value=tech).border = thin_border
        
        # Color Code
        fill_color = "FFFFFF"
        for k, v in colors.items():
            if k in tech: 
                fill_color = v
                break
        
        my_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        ws.cell(row=curr, column=COL_TECH).fill = my_fill
        
        # Timeline
        for d, count in gdata['days'].items():
            col_idx = COL_START_DAY + (d - 1)
            # Safety check for max columns
            if col_idx > 50: continue 
            
            cell = ws.cell(row=curr, column=col_idx)
            cell.value = count
            cell.fill = my_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            
        curr += 1
        
    print(f"[INFO] Total GPS Processed: {total_gps_verify}")
    wb.save(output_file)
    print(f"[SUCCESS] Saved to {output_file}")

if __name__ == "__main__":
    generate_gantt()
