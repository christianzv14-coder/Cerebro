
import openpyxl
import pandas as pd

def audit_gantt():
    fpath = "outputs/gantt_FINAL_RAW_314.xlsx"
    print(f"Auditing: {fpath}")
    
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        
        COL_TOTAL = 5 # E
        COL_START_DAY = 7 # G
        START_ROW = 5
        
        grand_total_col = 0
        grand_total_grid = 0
        
        row_count = 0
        
        for r in range(START_ROW, 500):
            # Check if row has data (City or Tech)
            city_val = ws.cell(row=r, column=4).value
            if not city_val:
                continue
                
            row_count += 1
            
            # Sum Column E (Total)
            t_val = ws.cell(row=r, column=COL_TOTAL).value
            if t_val:
                grand_total_col += float(t_val)
            
            # Sum Grid (G onwards)
            row_grid_sum = 0
            for c in range(COL_START_DAY, COL_START_DAY + 62):
                d_val = ws.cell(row=r, column=c).value
                if d_val is not None:
                    # Check if numeric
                    try:
                        val = float(d_val)
                        row_grid_sum += val
                        grand_total_grid += val
                    except:
                        pass # Ignore non-numeric (e.g. strings/dates)
            
            # Mismatch per row?
            if t_val and abs(float(t_val) - row_grid_sum) > 0.01:
                print(f"Row {r} Mismatch: Col Total={t_val} vs Grid Sum={row_grid_sum} (City: {city_val})")

        print(f"--- AUDIT RESULTS ---")
        print(f"Rows Found: {row_count}")
        print(f"Sum of 'Total' Column: {grand_total_col}")
        print(f"Sum of 'Daily Grid': {grand_total_grid}")
        
        if abs(grand_total_col - 314) < 0.1 and abs(grand_total_grid - 314) < 0.1:
            print("SUCCESS: Excel verifies to 314.")
        else:
            print("FAILURE: Excel does NOT sum to 314.")
            
    except Exception as e:
        print(f"Audit Failed: {e}")

if __name__ == "__main__":
    audit_gantt()
