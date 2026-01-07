
import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from datetime import datetime, timedelta

# --- CONFIGURATION ---
FILE_GANTT_UNIFIED = "temp_inspect.xlsx"
OUTPUT_FILE = "outputs/GANTT_UNIFICADO_V12_FINAL.xlsx"

START_DATE = datetime(2026, 1, 12)
END_DATE = datetime(2026, 1, 31)

def excel_date_to_datetime(val):
    try:
        if isinstance(val, (int, float)) and val > 44000:
            return datetime(1899, 12, 30) + timedelta(days=int(val))
        return pd.to_datetime(val, errors='coerce')
    except:
        return None

# WORKING DAYS
working_days_set = set()
curr = START_DATE
while curr <= END_DATE:
    if curr.weekday() != 6: # Monday to Saturday
        working_days_set.add(curr.strftime('%d-%m-%Y'))
    curr += timedelta(days=1)
ordered_working_days = sorted(list(working_days_set), key=lambda x: datetime.strptime(x, '%d-%m-%Y'))

# TECHNICIANS
tech_list = [
    {"name": "Luis", "base": "SANTIAGO"},
    {"name": "Efrain", "base": "SANTIAGO"},
    {"name": "Carlos", "base": "SANTIAGO"},
    {"name": "Wilmer", "base": "SANTIAGO"},
    {"name": "Fabian", "base": "SANTIAGO"},
    {"name": "Jimmy", "base": "CHILLAN"},
    {"name": "Orlando", "base": "CALAMA"}
]

TECH_NAMES = [t['name'].upper() for t in tech_list]

REGION_TO_CITY = {
    'METROPOLITANA DE SANTIAGO': 'SANTIAGO',
    'ANTOFAGASTA': 'ANTOFAGASTA',
    'ARICA': 'ARICA',
    'ATACAMA': 'COPIAPO',
    'COQUIMBO': 'LA SERENA',
    'LA ARAUCANÍA': 'TEMUCO',
    'LIBERTADOR GENERAL BERNARDO O\'HIGGINS': 'RANCAGUA',
    'MAULE': 'TALCA',
    'BIOBÍO': 'CONCEPCION',
    'ÑUBLE': 'CHILLAN',
    'LOS LAGOS': 'PUERTO MONTT',
    'VALPARAÍSO': 'VALPARAISO',
    'TARAPACÁ': 'IQUIQUE'
}

def load_data():
    xl = pd.ExcelFile(FILE_GANTT_UNIFIED)
    
    # 1. RABIE (Sheet 0)
    df_rabie_raw = xl.parse(0, header=None)
    date_cells = [] 
    for r in range(len(df_rabie_raw)):
        for c in range(len(df_rabie_raw.columns)):
            v = df_rabie_raw.iloc[r, c]
            d = excel_date_to_datetime(v)
            if pd.notna(d):
                d_str = d.strftime('%d-%m-%Y')
                if d_str in working_days_set:
                    date_cells.append((r, c, d_str))

    print(f"[DEBUG] Found {len(date_cells)} date cells in Rabie sheet.")

    rabie_demand = []
    for rd, cd, d_str in date_cells:
        for r in range(rd + 1, len(df_rabie_raw)):
            if any(dr == r and dc == cd for dr, dc, _ in date_cells): break
            val = df_rabie_raw.iloc[r, cd]
            try:
                q = float(val)
                if q > 0:
                    # City is strictly in Column 3 for Sheet 0 in the complex layout
                    city_val = str(df_rabie_raw.iloc[r, 3]).strip()
                    if city_val == 'nan' or city_val == '' or city_val.upper() == 'CIUDAD':
                        # Fallback crawler logic if Col 3 is empty
                        city = "UNKNOWN"
                        for c_search in range(cd - 1, -1, -1):
                            txt = str(df_rabie_raw.iloc[r, c_search]).strip().upper()
                            if txt != 'NAN' and txt != '' and not txt.replace('.','').isdigit() and txt not in TECH_NAMES:
                                city = txt
                                break
                    else:
                        city = city_val.upper()
                    
                    # If the crawler picked up a technician name, skip it or try harder
                    if city in TECH_NAMES:
                         # Try searching further left or right
                         for c_search in range(3, 10): # Most cities are in col 3 or 4
                             txt = str(df_rabie_raw.iloc[r, c_search]).strip().upper()
                             if txt != 'NAN' and txt != '' and not txt.replace('.','').isdigit() and txt not in TECH_NAMES:
                                 city = txt
                                 break

                    rabie_demand.append({'City': city, 'Day': d_str, 'Qty': int(q)})
            except: pass

    # 2. ENTEL (Sheet 1)
    df_entel_raw = xl.parse(1, header=None)
    entel_backlog = []
    # CSV Structure: Col 1 = Region, Col 2 = Qty, Col 3 = Priority
    for i in range(len(df_entel_raw)):
        row = df_entel_raw.iloc[i]
        reg_val = str(row[1]).strip().upper()
        if reg_val != 'NAN' and reg_val != '' and reg_val != 'REGIÓN':
            # EXCLUDE SANTIAGO FROM ENTEL BACKLOG
            if 'SANTIAGO' in reg_val or 'METROPOLITANA' in reg_val:
                continue

            try:
                qty = row[2]
                prio = row[3]
                if pd.notna(qty) and float(qty) > 0:
                    city = REGION_TO_CITY.get(reg_val, reg_val)
                    entel_backlog.append({
                        'City': city, 
                        'Qty': int(float(qty)), 
                        'Priority': int(float(prio)) if pd.notna(prio) else 3
                    })
            except: pass

    print(f"[DEBUG] Loaded {len(rabie_demand)} Rabie services and {sum(m['Qty'] for m in entel_backlog)} Entel backlog items.")
    return rabie_demand, entel_backlog

def get_dist(base, target_city, df_dist):
    if base == target_city: return 0
    try:
        # Distance matrix uses normalized names
        base_norm = base.upper().strip()
        target_norm = target_city.upper().strip()
        if base_norm in df_dist.index and target_norm in df_dist.columns:
            return df_dist.loc[base_norm, target_norm]
    except: pass
    return 9999

def plan_gantt():
    rabie_demand, entel_backlog = load_data()
    df_dist = pd.read_excel("data/matriz_distancia_km.xlsx", index_col=0)
    df_dist.columns = [str(c).upper().strip() for c in df_dist.columns]
    df_dist.index = [str(i).upper().strip() for i in df_dist.index]

    MAX_LOAD = 1
    assignments = {t['name']: {day: None for day in ordered_working_days} for t in tech_list}
    external_assignments = []

    # A. RABIE (P1)
    for item in rabie_demand:
        city, day, qty = item['City'], item['Day'], item['Qty']
        for _ in range(qty):
            assigned = False
            for t_info in sorted(tech_list, key=lambda t: get_dist(t['base'], city, df_dist)):
                t_name = t_info['name']
                if assignments[t_name][day] is None:
                    assignments[t_name][day] = {'city': city, 'type': 'R'}
                    assigned = True
                    break
            if not assigned:
                external_assignments.append({'City': city, 'Day': day, 'Type': 'R'})

    # B. ENTEL (P2-4)
    for item in sorted(entel_backlog, key=lambda x: x['Priority']):
        city, backlog = item['City'], item['Qty']
        for day in ordered_working_days:
            if backlog <= 0: break
            for t_info in sorted(tech_list, key=lambda t: get_dist(t['base'], city, df_dist)):
                if backlog <= 0: break
                t_name = t_info['name']
                if assignments[t_name][day] is None:
                    assignments[t_name][day] = {'city': city, 'type': 'E'}
                    backlog -= 1
        
        if backlog > 0:
            for i in range(backlog):
                day = ordered_working_days[i % len(ordered_working_days)]
                external_assignments.append({'City': city, 'Day': day, 'Type': 'E'})

    # --- RENDER ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GANTT UNIFICADO V12"
    fill_r = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type='solid') 
    fill_e = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type='solid') 
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(row=1, column=1, value="GANTT UNIFICADO V12 - SIN SANTIAGO ENTEL").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="REGLA: 1 unidad/día por técnico. [R]=Rabie, [E]=Entel Backlog.")
    
    row_idx = 4
    ws.cell(row=row_idx, column=1, value="Técnico").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value="Base").font = Font(bold=True)
    for i, day in enumerate(ordered_working_days):
        ws.cell(row=row_idx, column=3+i, value=day).alignment = Alignment(horizontal='center')

    row_idx += 1
    # Internals
    for t_info in tech_list:
        name = t_info['name']
        ws.cell(row=row_idx, column=1, value=name).border = border
        ws.cell(row=row_idx, column=2, value=t_info['base']).border = border
        for i, day in enumerate(ordered_working_days):
            it = assignments[name][day]
            cell = ws.cell(row=row_idx, column=3+i); cell.border = border
            if it:
                cell.value = f"{it['city']} [{it['type']}]"
                cell.alignment = Alignment(horizontal='center')
                cell.fill = fill_r if it['type'] == 'R' else fill_e
        row_idx += 1

    # Externals (Group by Day)
    ext_df = pd.DataFrame(external_assignments)
    if not ext_df.empty:
        max_ext = ext_df.groupby('Day').size().max()
        for j in range(max_ext):
            ws.cell(row=row_idx, column=1, value=f"Externo {j+1}").border = border
            ws.cell(row=row_idx, column=2, value="EXTERNO").border = border
            for i, day in enumerate(ordered_working_days):
                day_data = ext_df[ext_df['Day'] == day]
                cell = ws.cell(row=row_idx, column=3+i); cell.border = border
                if len(day_data) > j:
                    it = day_data.iloc[j]
                    cell.value = f"{it['City']} [{it['Type']}]"
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = fill_r if it['Type'] == 'R' else fill_e
            row_idx += 1

    wb.save(OUTPUT_FILE)
    print(f"[SUCCESS] V12 Final Generated. Total Assigned: {len(rabie_demand) + sum(m['Qty'] for m in entel_backlog)}")

if __name__ == "__main__":
    plan_gantt()

if __name__ == "__main__":
    plan_gantt()
