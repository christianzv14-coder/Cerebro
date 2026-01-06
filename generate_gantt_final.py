
import openpyxl
import json
import modelo_optimizacion_gps_chile_v1 as shared
from openpyxl.styles import PatternFill, Alignment, Border, Side

def generate_gantt():
    # 1. Load Data
    with open("outputs/vrp_result.json", "r") as f:
        data = json.load(f)
    plan = data['plan']
    
    # 2. Group by (City, Tech)
    grouped = {}
    
    for item in plan:
        c = item['city']
        t = item.get('tech', 'EXTERNAL')
        if item['type'] == 'EXTERNAL': t = 'External'
        
        key = (c, t)
        if key not in grouped:
            grouped[key] = {'total': 0, 'days': {}}
            
        grouped[key]['total'] += item['gps']
        
        # FIX: Accumulate daily counts (do not overwrite)
        d = item['day']
        current_day_val = grouped[key]['days'].get(d, 0)
        grouped[key]['days'][d] = current_day_val + item['gps']
        
    sorted_keys = sorted(grouped.keys(), key=lambda x: (x[0], x[1]))
    
    # Debug: Check Antofagasta Grouping
    anto_key = ('Antofagasta', 'Orlando')
    if anto_key in grouped:
        print(f"DEBUG ANTOFAGASTA: {grouped[anto_key]}")
    else:
        # Try finding partial key
        for k in grouped:
            if 'Antofagasta' in k[0]:
                print(f"DEBUG ANTOFAGASTA KEY FOUND: {k} -> {grouped[k]}")
    
    # 3. Load Template
    src_file = "temp_gantt.xlsx"
    out_file = "outputs/gantt_Final_Revertido_FIX.xlsx"
    
    wb = openpyxl.load_workbook(src_file)
    ws = wb.active 
    
    # 4. Parameters (Aligned with Inspection & Screen 2)
    # Inspection: Col4=City, Col5=Total, Col6=Tech, Col7=Day1
    COL_CITY = 4 # D
    COL_TOTAL = 5 # E
    COL_TECH = 6 # F
    COL_START_DAY = 7 # G
    START_ROW = 5
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Colors
    colors = {
        "Carlos": "FFC7CE", "Luis": "C6EFCE", "Orlando": "FFEB9C",
        "Fabian": "BDD7EE", "Pedro": "E2EFDA", "Efrain": "FCE4D6",
        "Jimmy": "EDEDED", "Wilmer": "D9E1F2", "External": "A9A9A9"
    }
    
    # 5. Clear Old Data (Delete Rows for Clean Slate)
    # This removes any hidden merges or styles that conflict
    ws.delete_rows(START_ROW, 500)
             
    # 6. Write Rows
    curr = START_ROW
    total_written_gps = 0 # Initialize counter
    
    for (city, tech) in sorted_keys:
        gdata = grouped[(city, tech)]
        
        # City
        c_cell = ws.cell(row=curr, column=COL_CITY)
        c_cell.value = city
        c_cell.border = thin_border
        
        # Total
        t_cell = ws.cell(row=curr, column=COL_TOTAL)
        t_cell.value = gdata['total']
        t_cell.alignment = Alignment(horizontal='center')
        t_cell.border = thin_border
        
        # Tech
        tech_cell = ws.cell(row=curr, column=COL_TECH)
        tech_cell.value = tech
        tech_cell.border = thin_border
        
        # Grid
        fill = None
        for k, v in colors.items():
            if k in tech:
                fill = PatternFill(start_color=v, end_color=v, fill_type="solid")
                break
        
        for d, qty in gdata['days'].items():
            # Fix: Handle Day 0 (External) -> Map to Day 1
            if d == 0: d = 1
            
            # Write only if d > 0 (Sanity check)
            if d > 0:
                col_idx = COL_START_DAY + (d - 1)
                d_cell = ws.cell(row=curr, column=col_idx)
                
                # Overlap Fix: ACCUMULATE instead of overwrite
                current_val = d_cell.value if d_cell.value is not None else 0
                new_val = current_val + qty
                
                d_cell.value = new_val
                total_written_gps += qty # Accumulate total GPS written
                
                if fill:
                    d_cell.fill = fill
                d_cell.border = thin_border
                d_cell.alignment = Alignment(horizontal='center')
                
        curr += 1
        
    # Add Verification Row at Bottom
    verif_row = curr + 2
    ws.cell(row=verif_row, column=COL_CITY).value = "VERIFICACION TOTAL"
    ws.cell(row=verif_row, column=COL_CITY).font = openpyxl.styles.Font(bold=True)
    
    t_cell = ws.cell(row=verif_row, column=COL_TOTAL)
    t_cell.value = total_written_gps
    t_cell.font = openpyxl.styles.Font(bold=True, color="FF0000")
    t_cell.alignment = Alignment(horizontal='center')
    t_cell.border = thin_border

    wb.save("outputs/gantt_FINAL_VERIFICADO_314.xlsx")
    print(f"Gantt Fixed: outputs/gantt_FINAL_VERIFICADO_314.xlsx with {curr-START_ROW} rows.")
    print(f"Total GPS Written to Gantt: {total_written_gps}") # Print total GPS written
    
    # Verification Sum
    print("--- VERIFICATION ---")
    
    total_written = 0
    for key, data in grouped.items():
        total_written += data['total']
    print(f"Total GPS in Data Source: {total_written}")

if __name__ == "__main__":
    generate_gantt()
