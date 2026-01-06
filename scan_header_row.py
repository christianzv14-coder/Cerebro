
import openpyxl

def scan_header():
    fpath = "c:/Users/chzam/OneDrive/Desktop/cerebro-patio/Cerebro/Detalle Entel.xlsx"
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb['Sep'] # Start with Sep
    
    print("Scanning Row 1 (Headers) up to Col 100...")
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=100, values_only=True):
        for i, val in enumerate(row):
            print(f"Col {i+1}: {val}")
            
    print("\nScanning Row 2 (Data) up to Col 100...")
    for row in ws.iter_rows(min_row=2, max_row=2, max_col=100, values_only=True):
        for i, val in enumerate(row):
            if val is not None:
                print(f"Col {i+1}: {val}")

if __name__ == "__main__":
    scan_header()
