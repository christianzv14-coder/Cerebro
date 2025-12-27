import openpyxl
import os

def inspect_excel():
    print("--- INSPECTOR GADGET ---")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    source_file = os.path.join(script_dir, "plantilla_planificacion_v2.xlsx")
    
    print(f"File: {source_file}")
    if not os.path.exists(source_file):
        print("FILE NOT FOUND!")
        return

    try:
        wb = openpyxl.load_workbook(source_file, data_only=True)
        print(f"Sheets found: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- SHEET: {sheet_name} ---")
            
            # Read header
            headers = [cell.value for cell in ws[1]]
            print(f"Headers: {headers}")
            
            # Read first 10 rows
            print("Content (First 10 rows):")
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True)):
                # Filter useful columns
                # finding index of 'tecnico_nombre'
                try:
                    tech_idx = headers.index('tecnico_nombre')
                    ticket_idx = headers.index('ticket_id')
                    print(f"Row {i+2}: Ticket={row[ticket_idx]} | Tech={row[tech_idx]}")
                except:
                    # Fallback print all
                    print(f"Row {i+2}: {row}")
                    
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_excel()
