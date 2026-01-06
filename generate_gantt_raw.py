
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import json

def generate_gantt_raw():
    # 1. Load Data
    with open("outputs/vrp_result.json", "r") as f:
        data = json.load(f)
    plan = data['plan']
    
    # 2. Group by (City, Tech)
    grouped = {}
    
    for item in plan:
        c = item['city']
        t = item.get('tech', 'EXTERNAL')
        if item['type'] == 'EXTERNAL': t = 'External'
        
        key = (c, t)
        if key not in grouped:
            grouped[key] = {'total': 0, 'days': {}}
            
        grouped[key]['total'] += item['gps']
        
        d = item['day']
        current_day_val = grouped[key]['days'].get(d, 0)
        grouped[key]['days'][d] = current_day_val + item['gps']
        
    sorted_keys = sorted(grouped.keys(), key=lambda x: (x[0], x[1]))
    
    # 3. Create Fresh Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Carta Gantt"
    
    # Styles
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    bold_font = Font(bold=True)
    
    colors = {
        "Fabian D.": "FFCCCB", "Fabian S.": "90EE90", "Orlando": "ADD8E6", 
        "Cristobal": "FFFFE0", "Carlos": "E6E6FA", "Felipe": "FFDAB9", 
        "Jimmy": "EDEDED", "Wilmer": "D9E1F2", "External": "A9A9A9"
    }

    # 4. Write Headers
    ws.cell(row=4, column=1).value = "Ciudad"
    ws.cell(row=4, column=2).value = "Técnico"
    ws.cell(row=4, column=3).value = "Total GPS"
    
    for c in range(1, 4):
        cell = ws.cell(row=4, column=c)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        
    # Day Headers (1 to 62)
    COL_START_DAY = 4
    for d in range(1, 63):
        cell = ws.cell(row=4, column=COL_START_DAY + d - 1)
        cell.value = d
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        # Set column width small
        col_letter = openpyxl.utils.get_column_letter(COL_START_DAY + d - 1)
        ws.column_dimensions[col_letter].width = 4

    # 5. Write Data Rows
    curr = 5
    total_written_gps = 0
    
    for (city, tech) in sorted_keys:
        gdata = grouped[(city, tech)]
        
        # City
        ws.cell(row=curr, column=1).value = city
        ws.cell(row=curr, column=1).border = thin_border
        
        # Tech
        ws.cell(row=curr, column=2).value = tech
        ws.cell(row=curr, column=2).border = thin_border
        
        # Total
        ws.cell(row=curr, column=3).value = gdata['total']
        ws.cell(row=curr, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=curr, column=3).border = thin_border
        
        # Color
        fill = None
        for k, v in colors.items():
            if k in tech:
                fill = PatternFill(start_color=v, end_color=v, fill_type="solid")
                break
        
        # Days
        for d, qty in gdata['days'].items():
            if d == 0: d = 1
            if d > 0:
                col_idx = COL_START_DAY + (d - 1)
                d_cell = ws.cell(row=curr, column=col_idx)
                
                # Accumulate logic not strictly needed if we iterate keys once, 
                # but good for safety if day 0 maps to 1
                current_val = d_cell.value if d_cell.value is not None else 0
                new_val = current_val + qty
                d_cell.value = new_val
                
                total_written_gps += qty
                
                if fill: d_cell.fill = fill
                d_cell.border = thin_border
                d_cell.alignment = Alignment(horizontal='center')
                
        curr += 1
        
    # Verification Row
    ws.cell(row=curr+1, column=1).value = "TOTAL VERIFICADO"
    ws.cell(row=curr+1, column=1).font = bold_font
    
    ws.cell(row=curr+1, column=3).value = total_written_gps
    ws.cell(row=curr+1, column=3).font = Font(bold=True, size=12, color="FF0000")
    ws.cell(row=curr+1, column=3).border = thin_border
    
    # Auto-width for City/Tech
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    out_file = "outputs/gantt_FINAL_RAW_314.xlsx"
    wb.save(out_file)
    print(f"Raw Gantt Generated: {out_file}")
    print(f"Total Written: {total_written_gps}")

if __name__ == "__main__":
    generate_gantt_raw()
